import os
import sys
import json
import ast
import re
from typing import Dict, Optional, List

# =========================
# 内联依赖检查（避免导入问题）
# =========================

# 依赖检查和备用方案
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
    print("✓ pandas已安装")
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️ pandas未安装，将使用基础Python功能")
    
    # 创建简单的pandas替代
    class SimpleDataFrame:
        def __init__(self, data=None, columns=None):
            self.data = data or []
            self.columns = columns or []
            
        def __len__(self):
            return len(self.data)
            
        def to_excel(self, filename, **kwargs):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # 写入列名
                if self.columns:
                    for col, header in enumerate(self.columns, 1):
                        ws.cell(row=1, column=col, value=header)
                
                # 写入数据
                for row_idx, row_data in enumerate(self.data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(filename)
                print(f"✓ 数据已保存到: {filename}")
            except ImportError:
                print("❌ openpyxl未安装，无法保存Excel文件")
                # 保存为CSV作为备用
                with open(filename.replace('.xlsx', '.csv'), 'w', encoding='utf-8') as f:
                    if self.columns:
                        f.write(','.join(str(col) for col in self.columns) + '\n')
                    for row_data in self.data:
                        f.write(','.join(str(val) for val in row_data) + '\n')
                print(f"✓ 数据已保存为CSV: {filename.replace('.xlsx', '.csv')}")
    
    # 创建pandas别名
    pd = type('pandas', (), {
        'DataFrame': SimpleDataFrame,
        'isna': lambda x: x is None or (hasattr(x, '__float__') and str(x) == 'nan'),
        'read_excel': lambda *args, **kwargs: SimpleDataFrame()
    })()

try:
    import numpy as np
    NUMPY_AVAILABLE = True
    print("✓ numpy已安装")
except ImportError:
    NUMPY_AVAILABLE = False
    print("⚠️ numpy未安装，将使用基础Python功能")
    
    # 创建numpy别名
    np = type('numpy', (), {
        'mean': lambda x: sum(x) / len(x) if x else 0,
        'std': lambda x: 0,  # 简化版本
        'nan': float('nan')
    })()

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
    print("✓ openpyxl已安装")
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl未安装，Excel功能受限")
    
    # 创建openpyxl备用对象
    class SimpleWorkbook:
        def __init__(self):
            self.active = SimpleWorksheet()
            
    class SimpleWorksheet:
        def __init__(self):
            self.cells = {}
            
        def cell(self, row, column, value=None):
            if value is not None:
                self.cells[(row, column)] = value
            return SimpleCell()
            
    class SimpleCell:
        def __init__(self):
            pass
            
    openpyxl = type('openpyxl', (), {
        'Workbook': SimpleWorkbook
    })()

# =========================
# 通用配置（按你的新评分结构）
# =========================

DEFAULT_ORDER = [
    "safety_compliance",
    "intent_understanding",
    "relevance",
    "accuracy",
    "comprehensiveness",
    "credibility",
    "timeliness",
    "logical_expression",
    "efficiency",
]

START_COL_1_BASED = 4  # 从第4列开始求平均（1基）
PREFER_ID_NAME = "序号"  # 优先作为分组的列名

# =========================
# 工具与解析函数
# =========================

def _to_number(v):
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None

def extract_k1(val):
    # 支持 dict、JSON 字符串、单引号 Python 字面量
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, dict):
        return _to_number(val.get("k1"))
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        # 尝试 JSON
        try:
            obj = json.loads(s)
            if isinstance(obj, dict):
                return _to_number(obj.get("k1"))
        except Exception:
            pass
        # 尝试 Python 字面量
        try:
            obj = ast.literal_eval(s)
            if isinstance(obj, dict):
                return _to_number(obj.get("k1"))
        except Exception:
            pass
    return None

# 主函数
def main():
    print("问答提取脚本启动成功！")
    print(f"pandas可用: {PANDAS_AVAILABLE}")
    print(f"numpy可用: {NUMPY_AVAILABLE}")
    print(f"openpyxl可用: {OPENPYXL_AVAILABLE}")
    
    # 查找Excel文件
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    if excel_files:
        print(f"找到Excel文件: {excel_files}")
        # 这里可以添加您的处理逻辑
    else:
        print("未找到Excel文件")

if __name__ == "__main__":
    main()
