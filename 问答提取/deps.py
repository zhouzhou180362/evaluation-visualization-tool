"""
问答提取模块的依赖检查
"""

import sys
import os

# 依赖检查和备用方案
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
    print("✓ pandas已安装")
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️ pandas未安装，将使用基础Python功能")
    
    # 创建简单的pandas替代
    class SimpleDataFrame:
        def __init__(self, data=None, columns=None):
            self.data = data or []
            self.columns = columns or []
            
        def __len__(self):
            return len(self.data)
            
        def to_excel(self, filename, **kwargs):
            try:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # 写入列名
                if self.columns:
                    for col, header in enumerate(self.columns, 1):
                        ws.cell(row=1, column=col, value=header)
                
                # 写入数据
                for row_idx, row_data in enumerate(self.data, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                wb.save(filename)
                print(f"✓ 数据已保存到: {filename}")
            except ImportError:
                print("❌ openpyxl未安装，无法保存Excel文件")
                # 保存为CSV作为备用
                with open(filename.replace('.xlsx', '.csv'), 'w', encoding='utf-8') as f:
                    if self.columns:
                        f.write(','.join(str(col) for col in self.columns) + '\n')
                    for row_data in self.data:
                        f.write(','.join(str(val) for val in row_data) + '\n')
                print(f"✓ 数据已保存为CSV: {filename.replace('.xlsx', '.csv')}")
    
    # 创建pandas别名
    pd = type('pandas', (), {
        'DataFrame': SimpleDataFrame,
        'isna': lambda x: x is None or (hasattr(x, '__float__') and str(x) == 'nan'),
        'read_excel': lambda *args, **kwargs: SimpleDataFrame()
    })()

try:
    import numpy as np
    NUMPY_AVAILABLE = True
    print("✓ numpy已安装")
except ImportError:
    NUMPY_AVAILABLE = False
    print("⚠️ numpy未安装，将使用基础Python功能")
    
    # 创建numpy别名
    np = type('numpy', (), {
        'mean': lambda x: sum(x) / len(x) if x else 0,
        'std': lambda x: 0,  # 简化版本
        'nan': float('nan')
    })()

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
    print("✓ openpyxl已安装")
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl未安装，Excel功能受限")

# 导出所有模块
__all__ = ['pd', 'np', 'openpyxl', 'PANDAS_AVAILABLE', 'NUMPY_AVAILABLE', 'OPENPYXL_AVAILABLE']
